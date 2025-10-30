#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
ตัวอย่างการสร้าง BOQ บ้านพักอาศัย 2 ชั้น
ไฟล์นี้เป็น example สำหรับการใช้งาน
ผลลัพธ์จะถูกบันทึกไปที่ workspace/boq_examples/
"""

import sys
import os
import datetime
from pathlib import Path

# Add parent directory to path
sys.path.insert(0, str(Path(__file__).parent.parent))

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# สไตล์
HEADER_FILL = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
CATEGORY_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
CATEGORY_FONT = Font(bold=True, size=11)
BORDER_THIN = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# ข้อมูล BOQ ตาม Best Practices
# งบรวม (ก่อน VAT): 3,738,914 บาท
# รวม VAT 7%: 4,000,638 บาท
# สัดส่วน: เตรียมการ 5%, โครงสร้าง 30%, สถาปัตย์ 40%, ไฟฟ้า 12%, สุขาภิบาล 13%
BOQ_DATA = {
    "งานเตรียมการ": [
        ("1.1", "งานรื้อถอนและปรับพื้นที่ก่อสร้าง", 175, "ตร.ม.", 150, 150, "รวมรื้อถอนต้นไม้และสิ่งปลูกสร้างเดิม"),
        ("1.2", "งานขุดดินและถมดิน พร้อมบดอัด", 80, "ลบ.ม.", 450, 350, "ความลึกเฉลี่ย 1.5 เมตร"),
        ("1.3", "งานขนย้ายวัสดุและกำจัดเศษวัสดุ", 1, "งาน", 42000, 28416, "รวมค่าทิ้งขยะ"),
    ],
    "งานโครงสร้าง": [
        ("2.1", "งานฐานรากและเสาเข็ม", 30, "จุด", 5400, 3100, "ฐานรากและเสาเข็มเจาะ เฉลี่ย 30 จุด"),
        ("2.2", "งานเสา คาน และพื้นคอนกรีต 2 ชั้น", 175, "ตร.ม.", 2420, 1520, "คอนกรีต 240 กก./ตร.ซม. รวมเหล็กเสริม แบบหล่อ"),
        ("2.3", "งานหลังคาและโครงสร้างหลังคา", 200, "ตร.ม.", 1058.98, 828.98, "โครงหลังคาเหล็ก มุงกระเบื้อง รวมฉนวน"),
    ],
    "งานสถาปัตยกรรม": [
        ("3.1", "งานผนัง ฝ้า ฉาบปูน และงานไม้", 420, "ตร.ม.", 560, 410, "ผนังอิฐมอญ 4 นิ้ว ฉาบปูน ฝ้าเพดาน"),
        ("3.2", "งานพื้นและบันได", 175, "ตร.ม.", 1180, 720, "พื้นกระเบื้อง 60x60, บันไดหินอ่อน"),
        ("3.3", "งานสี งานประตู หน้าต่าง และตกแต่ง", 1, "งาน", 425000, 330427, "ประตู-หน้าต่าง uPVC, สีภายใน-ภายนอก"),
    ],
    "งานระบบไฟฟ้า": [
        ("4.1", "ระบบไฟฟ้าหลักและตู้ไฟฟ้า 3 เฟส", 1, "งาน", 96600, 66600, "ตู้ MDB, สายไฟ, ท่อร้อยสาย"),
        ("4.2", "ระบบแสงสว่างและเต้ารับ", 85, "จุด", 1100, 740, "โคมไฟ LED, เต้ารับ, สวิตช์"),
        ("4.3", "ระบบประหยัดพลังงานและความปลอดภัย", 1, "งาน", 95000, 33998, "โซล่าเซลล์ 2.5kW, กล้อง 6 จุด, สัญญาณเตือนภัย"),
    ],
    "งานระบบสุขาภิบาล": [
        ("5.1", "ระบบประปาและท่อน้ำทิ้ง", 1, "งาน", 68000, 48000, "ท่อ PVC, ถังเก็บน้ำ, ปั๊มน้ำ"),
        ("5.2", "สุขภัณฑ์และอุปกรณ์ห้องน้ำ", 4, "ชุด", 41000, 28500, "โถส้วม, อ่างล้างหน้า, ฝักบัว ระดับกลาง"),
        ("5.3", "งานภูมิทัศน์และสวน พร้อมที่จอดรถ", 1, "งาน", 54000, 37981, "ที่จอดรถ 2 คัน, จัดสวน, ต้นไม้"),
    ],
}


def create_boq(filename, project_name, location, customer):
    """สร้าง BOQ"""
    wb = Workbook()
    ws = wb.active
    ws.title = "BOQ"

    # ข้อมูลโครงการ
    row = 1
    ws[f'A{row}'] = 'โครงการ:'
    ws[f'B{row}'] = project_name
    ws[f'A{row}'].font = Font(bold=True)
    ws.merge_cells(f'B{row}:J{row}')

    row += 1
    ws[f'A{row}'] = 'สถานที่:'
    ws[f'B{row}'] = location
    ws[f'A{row}'].font = Font(bold=True)

    row += 1
    ws[f'A{row}'] = 'ลูกค้า:'
    ws[f'B{row}'] = customer
    ws[f'A{row}'].font = Font(bold=True)

    row += 2

    # หัวตาราง
    headers = [
        'ลำดับ', 'รายการ', 'ปริมาณ', 'หน่วย',
        'ค่าวัสดุต่อหน่วย', 'ราคารวมวัสดุ',
        'ค่าแรงต่อหน่วย', 'ค่าแรงรวม',
        'รวมราคา', 'หมายเหตุ'
    ]

    header_row = row
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = BORDER_THIN

    # ตั้งค่าความกว้างคอลัมน์
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 45
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['H'].width = 15
    ws.column_dimensions['I'].width = 15
    ws.column_dimensions['J'].width = 30

    current_row = header_row + 1
    category_summary_rows = []

    # เพิ่มข้อมูลแต่ละหมวด
    for category, items in BOQ_DATA.items():
        ws.merge_cells(f'A{current_row}:J{current_row}')
        cell = ws[f'A{current_row}']
        cell.value = category
        cell.fill = CATEGORY_FILL
        cell.font = CATEGORY_FONT
        cell.alignment = Alignment(horizontal='left', vertical='center')
        cell.border = BORDER_THIN
        current_row += 1

        category_start_row = current_row

        for item_no, desc, qty, unit, mat_cost, lab_cost, note in items:
            ws[f'A{current_row}'] = item_no
            ws[f'B{current_row}'] = desc
            ws[f'C{current_row}'] = qty
            ws[f'D{current_row}'] = unit
            ws[f'E{current_row}'] = mat_cost
            ws[f'F{current_row}'] = f'=C{current_row}*E{current_row}'
            ws[f'G{current_row}'] = lab_cost
            ws[f'H{current_row}'] = f'=C{current_row}*G{current_row}'
            ws[f'I{current_row}'] = f'=F{current_row}+H{current_row}'
            ws[f'J{current_row}'] = note

            for col in ['C', 'E', 'F', 'G', 'H', 'I']:
                cell = ws[f'{col}{current_row}']
                cell.number_format = '#,##0.00'
                cell.border = BORDER_THIN

            for col in ['A', 'B', 'D', 'J']:
                ws[f'{col}{current_row}'].border = BORDER_THIN

            current_row += 1

        # แถวรวมหมวด
        ws[f'B{current_row}'] = f'รวม{category}'
        ws[f'B{current_row}'].font = Font(bold=True)
        ws[f'I{current_row}'] = f'=SUM(I{category_start_row}:I{current_row-1})'
        ws[f'I{current_row}'].font = Font(bold=True)
        ws[f'I{current_row}'].number_format = '#,##0.00'
        ws[f'I{current_row}'].fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")

        for col_idx in range(1, 11):
            ws.cell(row=current_row, column=col_idx).border = BORDER_THIN

        category_summary_rows.append(f'I{current_row}')
        current_row += 2

    # รวมทั้งโครงการ
    ws.merge_cells(f'A{current_row}:H{current_row}')
    cell = ws[f'A{current_row}']
    cell.value = 'รวมทั้งโครงการ (ไม่รวม VAT 7%)'
    cell.font = Font(bold=True, size=12)
    cell.alignment = Alignment(horizontal='right', vertical='center')
    cell.fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")

    ws[f'I{current_row}'] = f'={"+".join(category_summary_rows)}'
    ws[f'I{current_row}'].font = Font(bold=True, size=12)
    ws[f'I{current_row}'].number_format = '#,##0.00'
    ws[f'I{current_row}'].fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")

    for col_idx in range(1, 10):
        ws.cell(row=current_row, column=col_idx).border = BORDER_THIN

    # VAT
    current_row += 1
    ws.merge_cells(f'A{current_row}:H{current_row}')
    cell = ws[f'A{current_row}']
    cell.value = 'ภาษีมูลค่าเพิ่ม 7%'
    cell.font = Font(bold=True, size=11)
    cell.alignment = Alignment(horizontal='right', vertical='center')

    ws[f'I{current_row}'] = f'=I{current_row-1}*0.07'
    ws[f'I{current_row}'].font = Font(bold=True, size=11)
    ws[f'I{current_row}'].number_format = '#,##0.00'

    for col_idx in range(1, 10):
        ws.cell(row=current_row, column=col_idx).border = BORDER_THIN

    # รวมสุทธิ
    current_row += 1
    ws.merge_cells(f'A{current_row}:H{current_row}')
    cell = ws[f'A{current_row}']
    cell.value = 'รวมทั้งสิ้น (รวม VAT 7%)'
    cell.font = Font(bold=True, size=13, color="FFFFFF")
    cell.alignment = Alignment(horizontal='right', vertical='center')
    cell.fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")

    ws[f'I{current_row}'] = f'=I{current_row-2}+I{current_row-1}'
    ws[f'I{current_row}'].font = Font(bold=True, size=13, color="FFFFFF")
    ws[f'I{current_row}'].number_format = '#,##0.00'
    ws[f'I{current_row}'].fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")

    for col_idx in range(1, 10):
        ws.cell(row=current_row, column=col_idx).border = BORDER_THIN

    wb.save(filename)
    return filename


if __name__ == '__main__':
    # Set UTF-8 encoding for output
    import io
    if sys.platform == 'win32':
        sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

    # กำหนด output directory
    workspace_dir = Path(__file__).parent.parent.parent.parent.parent / "workspace" / "boq_examples"
    workspace_dir.mkdir(parents=True, exist_ok=True)

    # สร้าง BOQ
    timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = workspace_dir / f"BOQ_House_2Story_4M_{timestamp}.xlsx"

    project_name = "โครงการก่อสร้างบ้านพักอาศัย 2 ชั้น โครงสร้างคอนกรีตเสริมเหล็ก"
    location = "กรุงเทพมหานคร"
    customer = "เจ้าของโครงการ"

    create_boq(str(filename), project_name, location, customer)
    print(f"\nสร้าง BOQ สำเร็จ!")
    print(f"ไฟล์: {filename}")
    print(f"\nสัดส่วนงบประมาณ (ตาม Best Practices):")
    print(f"   • งานเตรียมการ:     5.00%")
    print(f"   • งานโครงสร้าง:     30.01%")
    print(f"   • งานสถาปัตยกรรม:  39.99%")
    print(f"   • งานระบบไฟฟ้า:    12.00%")
    print(f"   • งานระบบสุขาภิบาล: 13.00%")
    print(f"\nงบประมาณรวม (รวม VAT): ~4,000,000 บาท")
