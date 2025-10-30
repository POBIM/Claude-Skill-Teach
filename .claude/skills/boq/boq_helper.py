#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
BOQ (Bill of Quantities) Helper Script
สร้างและจัดการใบประมาณราคาค่าก่อสร้าง

ไฟล์ BOQ ที่สร้างจะถูกบันทึกไปที่ workspace/boq_examples/
เพื่อแยกออกจาก skill folder
"""

import sys
import json
import io
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from pathlib import Path

# Set UTF-8 encoding for Windows
if sys.platform == 'win32':
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8')


# หมวดงานมาตรฐาน
WORK_CATEGORIES = [
    "งานเตรียมการ",
    "งานโครงสร้าง",
    "งานสถาปัตยกรรม",
    "งานระบบไฟฟ้า",
    "งานระบบสุขาภิบาล"
]

# สไตล์มาตรฐาน
HEADER_FILL = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
CATEGORY_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
CATEGORY_FONT = Font(bold=True, size=11)
BORDER_THIN = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)


def create_boq(filename, project_name="", location="", customer=""):
    """สร้าง BOQ ใหม่"""
    wb = Workbook()
    ws = wb.active
    ws.title = "BOQ"

    # ข้อมูลโครงการ
    row = 1
    ws[f'A{row}'] = 'โครงการ:'
    ws[f'B{row}'] = project_name
    ws[f'A{row}'].font = Font(bold=True)

    row += 1
    ws[f'A{row}'] = 'สถานที่:'
    ws[f'B{row}'] = location
    ws[f'A{row}'].font = Font(bold=True)

    row += 1
    ws[f'A{row}'] = 'ลูกค้า:'
    ws[f'B{row}'] = customer
    ws[f'A{row}'].font = Font(bold=True)

    row += 2

    # หัวตาราง
    headers = [
        'ลำดับ', 'รายการ', 'ปริมาณ', 'หน่วย',
        'ค่าวัสดุต่อหน่วย', 'ราคารวมวัสดุ',
        'ค่าแรงต่อหน่วย', 'ค่าแรงรวม',
        'รวมราคา', 'หมายเหตุ'
    ]

    header_row = row
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = BORDER_THIN

    # ตั้งค่าความกว้างคอลัมน์
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['H'].width = 15
    ws.column_dimensions['I'].width = 15
    ws.column_dimensions['J'].width = 20

    current_row = header_row + 1

    # เพิ่มหมวดงานทั้งหมด
    for category in WORK_CATEGORIES:
        # แถวหมวดงาน
        ws.merge_cells(f'A{current_row}:J{current_row}')
        cell = ws[f'A{current_row}']
        cell.value = category
        cell.fill = CATEGORY_FILL
        cell.font = CATEGORY_FONT
        cell.alignment = Alignment(horizontal='left', vertical='center')
        cell.border = BORDER_THIN
        current_row += 1

        # เว้นแถวสำหรับรายการ (3 แถว)
        for _ in range(3):
            for col_idx in range(1, 11):
                cell = ws.cell(row=current_row, column=col_idx)
                cell.border = BORDER_THIN
                if col_idx in [3, 5, 7]:  # ปริมาณ, ค่าวัสดุ, ค่าแรง
                    cell.number_format = '#,##0.00'
                elif col_idx in [6, 8, 9]:  # ราคารวม
                    cell.number_format = '#,##0.00'
            current_row += 1

        # แถวรวมหมวด
        ws[f'A{current_row}'] = ''
        ws[f'B{current_row}'] = f'รวม{category}'
        ws[f'B{current_row}'].font = Font(bold=True)
        ws[f'I{current_row}'] = f'=SUM(I{current_row-3}:I{current_row-1})'
        ws[f'I{current_row}'].font = Font(bold=True)
        ws[f'I{current_row}'].number_format = '#,##0.00'

        for col_idx in range(1, 11):
            ws.cell(row=current_row, column=col_idx).border = BORDER_THIN

        current_row += 2

    # แถวรวมทั้งโครงการ
    ws.merge_cells(f'A{current_row}:H{current_row}')
    cell = ws[f'A{current_row}']
    cell.value = 'รวมทั้งโครงการ'
    cell.font = Font(bold=True, size=12)
    cell.alignment = Alignment(horizontal='right', vertical='center')
    cell.fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")

    # หาแถวรวมของแต่ละหมวด (ทุกๆ 6 แถว)
    summary_rows = []
    start_row = header_row + 1
    for i in range(len(WORK_CATEGORIES)):
        summary_row = start_row + (i * 6) + 4  # 1 แถวหมวด + 3 แถวรายการ + 1 แถวรวม
        summary_rows.append(f'I{summary_row}')

    ws[f'I{current_row}'] = f'={"+".join(summary_rows)}'
    ws[f'I{current_row}'].font = Font(bold=True, size=12)
    ws[f'I{current_row}'].number_format = '#,##0.00'
    ws[f'I{current_row}'].fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")

    for col_idx in range(1, 10):
        ws.cell(row=current_row, column=col_idx).border = BORDER_THIN

    wb.save(filename)
    return {
        'status': 'success',
        'message': f'สร้าง BOQ สำเร็จ: {filename}',
        'file': filename
    }


def add_item(filename, category, item_no, description, quantity, unit,
             material_cost=0, labor_cost=0, note=""):
    """เพิ่มรายการใน BOQ"""
    if not Path(filename).exists():
        return {'error': f'ไม่พบไฟล์ {filename}'}

    if category not in WORK_CATEGORIES:
        return {'error': f'หมวดงาน "{category}" ไม่ถูกต้อง'}

    wb = load_workbook(filename)
    ws = wb.active

    # หาแถวของหมวดงานที่ต้องการ
    category_idx = WORK_CATEGORIES.index(category)

    # หาแถวเริ่มต้น (หัวตาราง + 1)
    header_row = 6  # แถวหัวตาราง (หลังข้อมูลโครงการ 3 แถว + เว้น 2 แถว)
    start_row = header_row + 1 + (category_idx * 6) + 1  # +1 สำหรับแถวหมวด

    # หาแถวว่างในหมวดนี้
    target_row = None
    for i in range(3):
        row = start_row + i
        if ws[f'B{row}'].value is None:
            target_row = row
            break

    if target_row is None:
        wb.close()
        return {'error': f'หมวด "{category}" เต็มแล้ว (สูงสุด 3 รายการ)'}

    # เพิ่มข้อมูล
    ws[f'A{target_row}'] = item_no
    ws[f'B{target_row}'] = description
    ws[f'C{target_row}'] = quantity
    ws[f'D{target_row}'] = unit
    ws[f'E{target_row}'] = material_cost
    ws[f'F{target_row}'] = f'=C{target_row}*E{target_row}'
    ws[f'G{target_row}'] = labor_cost
    ws[f'H{target_row}'] = f'=C{target_row}*G{target_row}'
    ws[f'I{target_row}'] = f'=F{target_row}+H{target_row}'
    ws[f'J{target_row}'] = note

    # ฟอร์แมตตัวเลข
    for col in ['C', 'E', 'F', 'G', 'H', 'I']:
        ws[f'{col}{target_row}'].number_format = '#,##0.00'

    wb.save(filename)
    wb.close()

    return {
        'status': 'success',
        'message': f'เพิ่มรายการสำเร็จในแถว {target_row}',
        'row': target_row
    }


def main():
    if len(sys.argv) < 2:
        print(json.dumps({
            'error': 'กรุณาระบุคำสั่ง',
            'usage': {
                'create': 'python boq_helper.py create <filename> [project_name] [location] [customer]',
                'add': 'python boq_helper.py add <filename> <category> <item_no> <description> <quantity> <unit> [material_cost] [labor_cost] [note]'
            }
        }, ensure_ascii=False, indent=2))
        sys.exit(1)

    command = sys.argv[1]

    if command == 'create':
        if len(sys.argv) < 3:
            print(json.dumps({'error': 'กรุณาระบุชื่อไฟล์'}, ensure_ascii=False))
            sys.exit(1)

        filename = sys.argv[2]
        project_name = sys.argv[3] if len(sys.argv) > 3 else ""
        location = sys.argv[4] if len(sys.argv) > 4 else ""
        customer = sys.argv[5] if len(sys.argv) > 5 else ""

        result = create_boq(filename, project_name, location, customer)
        print(json.dumps(result, ensure_ascii=False, indent=2))

    elif command == 'add':
        if len(sys.argv) < 8:
            print(json.dumps({
                'error': 'ข้อมูลไม่ครบถ้วน',
                'required': ['filename', 'category', 'item_no', 'description', 'quantity', 'unit']
            }, ensure_ascii=False))
            sys.exit(1)

        filename = sys.argv[2]
        category = sys.argv[3]
        item_no = sys.argv[4]
        description = sys.argv[5]
        quantity = float(sys.argv[6])
        unit = sys.argv[7]
        material_cost = float(sys.argv[8]) if len(sys.argv) > 8 else 0
        labor_cost = float(sys.argv[9]) if len(sys.argv) > 9 else 0
        note = sys.argv[10] if len(sys.argv) > 10 else ""

        result = add_item(filename, category, item_no, description, quantity,
                         unit, material_cost, labor_cost, note)
        print(json.dumps(result, ensure_ascii=False, indent=2))

    else:
        print(json.dumps({
            'error': f'คำสั่ง "{command}" ไม่ถูกต้อง',
            'available_commands': ['create', 'add']
        }, ensure_ascii=False))
        sys.exit(1)


if __name__ == '__main__':
    main()
